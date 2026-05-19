#!/usr/bin/env python3
"""
Export Web3 (100k) and AI (133k+) research datasets to a multi-sheet Excel file.
Sheets:
  1. Web3 Documents  – all 100k records with title, source, url, date, category, type, citations, bigrams
  2. AI Documents    – all 133k+ records with the same columns
  3. Combined Stats  – keyword frequencies, bigram frequencies, year distributions, category breakdowns for both
"""

import json, re, os, sys
from collections import Counter
from itertools import islice

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Paths ──
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WEB3_PATH = os.path.join(BASE, "src", "lib", "web3-reports-data-generated.json")
AI_PATH   = os.path.join(BASE, "src", "lib", "ai-reports-data-generated.json")
OUTPUT    = os.path.join(BASE, "research-corpus-export.xlsx")

# ── Stopwords for bigram extraction ──
STOP = set("a an the and or but in on at to for of is it by as with from this that be are was were".split())

def extract_bigrams(title: str) -> str:
    """Extract meaningful bigrams from a title string."""
    if not title:
        return ""
    words = re.sub(r"[^a-z0-9\s]", "", title.lower()).split()
    words = [w for w in words if w not in STOP and len(w) > 1]
    bigrams = [f"{words[i]} {words[i+1]}" for i in range(len(words) - 1)]
    return "; ".join(bigrams[:5])  # top 5 bigrams per title

def generate_summary(title: str, category: str, doc_type: str, citations: int) -> str:
    """Generate a brief analytical summary from available metadata."""
    if not title:
        return ""
    parts = []
    parts.append(f"{doc_type} in {category}." if category else f"{doc_type}.")
    if citations and citations > 0:
        if citations > 1000:
            parts.append(f"Highly cited ({citations:,} citations).")
        elif citations > 100:
            parts.append(f"Well-cited ({citations:,} citations).")
        else:
            parts.append(f"{citations:,} citations.")
    # Extract key topic from title
    title_lower = title.lower()
    if any(kw in title_lower for kw in ["survey", "review", "overview"]):
        parts.append("Survey/review paper.")
    elif any(kw in title_lower for kw in ["framework", "architecture", "system"]):
        parts.append("Systems/framework contribution.")
    elif any(kw in title_lower for kw in ["analysis", "study", "evaluation"]):
        parts.append("Empirical analysis.")
    elif any(kw in title_lower for kw in ["novel", "new", "proposed"]):
        parts.append("Novel methodology.")
    return " ".join(parts)

def compute_stats(data: list, label: str) -> dict:
    """Compute keyword frequencies, bigram frequencies, year distributions, category breakdowns."""
    all_words = Counter()
    all_bigrams = Counter()
    year_dist = Counter()
    category_dist = Counter()
    type_dist = Counter()
    source_dist = Counter()
    citation_sum = 0
    citation_count = 0

    for doc in data:
        title = doc.get("title", "")
        if title:
            words = re.sub(r"[^a-z0-9\s]", "", title.lower()).split()
            words = [w for w in words if w not in STOP and len(w) > 2]
            all_words.update(words)
            for i in range(len(words) - 1):
                all_bigrams[f"{words[i]} {words[i+1]}"] += 1

        date = doc.get("date", "")
        if date:
            year = str(date)[:4]
            if year.isdigit():
                year_dist[year] += 1

        cat = doc.get("category", "")
        if cat:
            category_dist[cat] += 1

        dtype = doc.get("type", "")
        if dtype:
            type_dist[dtype] += 1

        src = doc.get("source", "")
        if src:
            source_dist[src] += 1

        cit = doc.get("citations", 0)
        if cit and isinstance(cit, (int, float)):
            citation_sum += cit
            citation_count += 1

    return {
        "label": label,
        "total": len(data),
        "keywords": all_words.most_common(200),
        "bigrams": all_bigrams.most_common(200),
        "years": sorted(year_dist.items()),
        "categories": category_dist.most_common(50),
        "types": type_dist.most_common(20),
        "sources": source_dist.most_common(20),
        "avg_citations": round(citation_sum / max(citation_count, 1), 1),
        "total_citations": citation_sum,
    }

# ── Styling ──
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
STAT_HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

def auto_width(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def main():
    print("Loading Web3 dataset...")
    with open(WEB3_PATH, "r") as f:
        web3_data = json.load(f)
    print(f"  → {len(web3_data):,} Web3 documents loaded")

    print("Loading AI dataset...")
    with open(AI_PATH, "r") as f:
        ai_data = json.load(f)
    print(f"  → {len(ai_data):,} AI documents loaded")

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════
    # Sheet 1: Web3 Documents
    # ══════════════════════════════════════════════
    print("Writing Web3 sheet...")
    ws_web3 = wb.active
    ws_web3.title = "Web3 Documents"
    headers = ["#", "Title", "Source", "URL", "Year", "Category", "Type", "Citations", "Bigrams", "Summary"]
    ws_web3.append(headers)
    style_header(ws_web3, len(headers))

    for i, doc in enumerate(web3_data, 1):
        title = doc.get("title", "")
        bigrams = extract_bigrams(title)
        summary = generate_summary(title, doc.get("category", ""), doc.get("type", ""), doc.get("citations", 0))
        ws_web3.append([
            i,
            title,
            doc.get("source", ""),
            doc.get("url", ""),
            str(doc.get("date", ""))[:4],
            doc.get("category", ""),
            doc.get("type", ""),
            doc.get("citations", 0),
            bigrams,
            summary,
        ])
        if i % 25000 == 0:
            print(f"  → {i:,} / {len(web3_data):,}")

    ws_web3.auto_filter.ref = f"A1:J{len(web3_data)+1}"
    ws_web3.freeze_panes = "B2"
    auto_width(ws_web3)

    # ══════════════════════════════════════════════
    # Sheet 2: AI Documents
    # ══════════════════════════════════════════════
    print("Writing AI sheet...")
    ws_ai = wb.create_sheet("AI Documents")
    ws_ai.append(headers)
    style_header(ws_ai, len(headers))

    for i, doc in enumerate(ai_data, 1):
        title = doc.get("title", "")
        bigrams = extract_bigrams(title)
        summary = generate_summary(title, doc.get("category", ""), doc.get("type", ""), doc.get("citations", 0))
        ws_ai.append([
            i,
            title,
            doc.get("source", ""),
            doc.get("url", ""),
            str(doc.get("date", ""))[:4],
            doc.get("category", ""),
            doc.get("type", ""),
            doc.get("citations", 0),
            bigrams,
            summary,
        ])
        if i % 25000 == 0:
            print(f"  → {i:,} / {len(ai_data):,}")

    ws_ai.auto_filter.ref = f"A1:J{len(ai_data)+1}"
    ws_ai.freeze_panes = "B2"
    auto_width(ws_ai)

    # ══════════════════════════════════════════════
    # Sheet 3: Combined Stats
    # ══════════════════════════════════════════════
    print("Computing statistics...")
    web3_stats = compute_stats(web3_data, "Web3")
    ai_stats = compute_stats(ai_data, "AI")

    ws_stats = wb.create_sheet("Combined Stats")
    row = 1

    # ── Overview ──
    def write_section(title_text, start_row):
        ws_stats.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
        cell = ws_stats.cell(row=start_row, column=1, value=title_text)
        cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(horizontal="left")
        return start_row + 1

    row = write_section("RESEARCH CORPUS OVERVIEW", row)
    overview_headers = ["Metric", "Web3", "AI", "Combined"]
    for col_idx, h in enumerate(overview_headers, 1):
        c = ws_stats.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = STAT_HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1

    overview_data = [
        ("Total Documents", f"{web3_stats['total']:,}", f"{ai_stats['total']:,}", f"{web3_stats['total'] + ai_stats['total']:,}"),
        ("Total Citations", f"{web3_stats['total_citations']:,}", f"{ai_stats['total_citations']:,}", f"{web3_stats['total_citations'] + ai_stats['total_citations']:,}"),
        ("Avg Citations/Doc", str(web3_stats['avg_citations']), str(ai_stats['avg_citations']), str(round((web3_stats['total_citations'] + ai_stats['total_citations']) / (web3_stats['total'] + ai_stats['total']), 1))),
        ("Unique Keywords (top 200)", "200", "200", "—"),
        ("Unique Bigrams (top 200)", "200", "200", "—"),
    ]
    for label, w3, ai_val, combined in overview_data:
        ws_stats.append([label, w3, ai_val, combined])
        row += 1

    row += 1

    # ── Top Keywords side by side ──
    row = write_section("TOP 100 KEYWORDS (by frequency)", row)
    kw_headers = ["Rank", "Web3 Keyword", "Web3 Count", "", "Rank", "AI Keyword", "AI Count"]
    for col_idx, h in enumerate(kw_headers, 1):
        c = ws_stats.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = STAT_HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1

    for i in range(100):
        w3_kw = web3_stats["keywords"][i] if i < len(web3_stats["keywords"]) else ("", "")
        ai_kw = ai_stats["keywords"][i] if i < len(ai_stats["keywords"]) else ("", "")
        ws_stats.append([i+1, w3_kw[0], w3_kw[1], "", i+1, ai_kw[0], ai_kw[1]])
        row += 1

    row += 1

    # ── Top Bigrams side by side ──
    row = write_section("TOP 100 BIGRAMS (by frequency)", row)
    bg_headers = ["Rank", "Web3 Bigram", "Web3 Count", "", "Rank", "AI Bigram", "AI Count"]
    for col_idx, h in enumerate(bg_headers, 1):
        c = ws_stats.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = STAT_HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1

    for i in range(100):
        w3_bg = web3_stats["bigrams"][i] if i < len(web3_stats["bigrams"]) else ("", "")
        ai_bg = ai_stats["bigrams"][i] if i < len(ai_stats["bigrams"]) else ("", "")
        ws_stats.append([i+1, w3_bg[0], w3_bg[1], "", i+1, ai_bg[0], ai_bg[1]])
        row += 1

    row += 1

    # ── Year Distribution ──
    row = write_section("PUBLICATION YEAR DISTRIBUTION", row)
    yr_headers = ["Year", "Web3 Count", "AI Count", "Combined"]
    for col_idx, h in enumerate(yr_headers, 1):
        c = ws_stats.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = STAT_HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1

    all_years = sorted(set(y for y, _ in web3_stats["years"]) | set(y for y, _ in ai_stats["years"]))
    w3_year_map = dict(web3_stats["years"])
    ai_year_map = dict(ai_stats["years"])
    for yr in all_years:
        w3c = w3_year_map.get(yr, 0)
        aic = ai_year_map.get(yr, 0)
        ws_stats.append([yr, w3c, aic, w3c + aic])
        row += 1

    row += 1

    # ── Category Breakdown ──
    row = write_section("CATEGORY BREAKDOWN", row)
    cat_headers = ["Web3 Category", "Count", "", "AI Category", "Count"]
    for col_idx, h in enumerate(cat_headers, 1):
        c = ws_stats.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = STAT_HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1

    max_cats = max(len(web3_stats["categories"]), len(ai_stats["categories"]))
    for i in range(max_cats):
        w3_cat = web3_stats["categories"][i] if i < len(web3_stats["categories"]) else ("", "")
        ai_cat = ai_stats["categories"][i] if i < len(ai_stats["categories"]) else ("", "")
        ws_stats.append([w3_cat[0], w3_cat[1], "", ai_cat[0], ai_cat[1]])
        row += 1

    row += 1

    # ── Source Breakdown ──
    row = write_section("DATA SOURCE BREAKDOWN", row)
    src_headers = ["Web3 Source", "Count", "", "AI Source", "Count"]
    for col_idx, h in enumerate(src_headers, 1):
        c = ws_stats.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = STAT_HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1

    max_srcs = max(len(web3_stats["sources"]), len(ai_stats["sources"]))
    for i in range(max_srcs):
        w3_src = web3_stats["sources"][i] if i < len(web3_stats["sources"]) else ("", "")
        ai_src = ai_stats["sources"][i] if i < len(ai_stats["sources"]) else ("", "")
        ws_stats.append([w3_src[0], w3_src[1], "", ai_src[0], ai_src[1]])
        row += 1

    auto_width(ws_stats)

    # ── Save ──
    print(f"Saving to {OUTPUT}...")
    wb.save(OUTPUT)
    print(f"✅ Done! File saved: {OUTPUT}")
    print(f"   → Sheet 1: Web3 Documents ({len(web3_data):,} rows)")
    print(f"   → Sheet 2: AI Documents ({len(ai_data):,} rows)")
    print(f"   → Sheet 3: Combined Stats (keywords, bigrams, years, categories, sources)")

if __name__ == "__main__":
    main()
